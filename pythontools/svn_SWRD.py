"""
This script is adapted from http://stackoverflow.com/a/10286163/2464295 but without the
multi threading factor of it.
Usage: `externals.py https://svn-server/repository`
Do note that the script is excluding src and hidden directories that starts with dot. Just modify the file
as it's not parameterized.
If you want this script to run faster, SSH to the SVN server and use file protocol.
Usage: `externals.py file:///svn/repository`
"""

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import svn_tools

def createSWRDxls(Path, resultFile):
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # call SVN to get externals
    p = svn_tools.analyzePath(Path)

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # summarize data
    worksheet_lines = list()
    for external in p:
        line = list()
        line.append(external[svn_tools.EXT_NAME])
        line.append(external[svn_tools.EXT_PATH])
        revision,author,date,msg = svn_tools.getLastCommitInfo(svn_tools.getAbsoluteSVNPath(external[svn_tools.EXT_PATH]))
        line.append(external[svn_tools.EXT_REVISION])
        line.append(msg)
        line.append(date)
        worksheet_lines.append(line)

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # generate Excel Workbook
    wb = Workbook()
    ws = wb.active # get active worksheet

    # header
    ws.append(["external name", Path, "revision"])

    # set column widths
    for i, width in enumerate([30,100,12,100,12]):
        ws.column_dimensions[get_column_letter(i+1)].width = width

    # write data to worksheet
    for line in worksheet_lines:
        ws.append(line)

    # activate Auto Filter
    ws.sheet_properties.filterMode = True

    bold = Font(bold=True)
    for cell in ['A1','B1','C1','D1','E1']:
        ws[cell].font=bold

    # highlight = Font(color="FF0000")
    # for row in ws.iter_rows(min_row=2):
    #     if (row[1].value != row[3].value):
    #         row[3].font = highlight
    #     if (row[2].value != row[4].value):
    #         row[4].font = highlight

    # Save the file
    wb.save(resultFile)

if __name__ == '__main__':
    Path = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/RefractiveWorkplace/RefWPClientInstaller/branches/releases/1.1"
    resultFile = "RefWP_branch_11.xlsx"
    createSWRDxls(Path, resultFile)
