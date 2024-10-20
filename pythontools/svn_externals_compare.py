"""
This script is adapted from http://stackoverflow.com/a/10286163/2464295 but without the
multi threading factor of it.
Usage: `externals.py https://svn-server/repository`
Do note that the script is excluding src and hidden directories that starts with dot. Just modify the file
as it's not parameterized.
If you want this script to run faster, SSH to the SVN server and use file protocol.
Usage: `externals.py file:///svn/repository`
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import sys
from xml.dom import minidom

EXT_PATH = "path"
EXT_REVISION = "revision"
EXT_NAME = "name"

def processLine(line):
    external = None
    content = line.split()
    if len(content)==2:
        external = dict()
        external[EXT_NAME] = content[1]
        t = content[0].partition('@')
        if len(t)==3:
            external[EXT_PATH] = t[0]
            if len(t[2])==0:
                external[EXT_REVISION] = "HEAD"
            else:
                external[EXT_REVISION] = t[2]
        else:
            external[EXT_PATH]=content[0]
            external[EXT_REVISION] = "HEAD"
        #print(">>>" + str(content))
        #print(">>> external: {} --- {}".format(content[0], content[1]))
    return external

def getAbsoluteSVNPath(url):
    #print("SVN relative path:" + url)
    absPath = url
    if (url.startswith(r"^/..")):
        absPath = url.replace(r"^/..",r"https://gullstrand.zeiss.org/svn",1)
    elif (url.startswith(r"^/")):
        absPath = url.replace(r"^/",r"https://gullstrand.zeiss.org/svn/czm/",1)
    #print("SVN absolute path:" + absPath)
    return absPath

def getExterns(url, base_dir):
    result = None
    print(">> external of {} (base dir: '{}')".format(url, base_dir))

    cmd = 'svn propget svn:externals "%s"' % url
    pipe = os.popen(cmd, 'r')
    data = pipe.read()
    pipe.close()

    if (data):
        print("#", url)
        #print(data.strip())
        result = list()
        for line in data.splitlines():
            external = processLine(line)
            if external:
                result.append(external)

    return result

def getLastCommitInfo(url):
    cmd = 'svn log -l 1 --xml "%s"' % url
    pipe = os.popen(cmd, 'r')
    data = pipe.read()
    pipe.close()

    # parse xml output of svn command
    mydoc = minidom.parseString(data)

    revisionElements = mydoc.getElementsByTagName('logentry')
    revision = revisionElements[0].attributes['revision'].value
    #print('revision:' + revision)

    authorElements = mydoc.getElementsByTagName('author')
    author = authorElements[0].firstChild.data
    #print('author:' + author)

    dateElements = mydoc.getElementsByTagName('date')
    date = dateElements[0].firstChild.data
    #print('date:' + date)

    msgElements = mydoc.getElementsByTagName('msg')
    msg = msgElements[0].firstChild.data
    #print('msg:' + msg)
    return revision,author,date,msg

def processDir(url, base_dir, recursive=False):
    url = getAbsoluteSVNPath(url)
    result = getExterns(url, base_dir )

    if (recursive):
        cmd = 'svn list "%s"' % url
        pipe = os.popen(cmd, 'r')
        listing = pipe.read().splitlines()
        pipe.close()

        dir_list = []
        for node in listing:
            if node.endswith('/') and not node.endswith('src/') and not node.startswith('.'):
                dir_list.append(node)

        for node in dir_list:
            print("> url:'{}', node:'{}', base_dir:'{}'".format(url, node, base_dir))
            processDir(url+node, base_dir+node )

    return result

def analyzePath(url, base_dir = ''):
    result = processDir(url, base_dir, recursive=False )
    print(result)
    return result

def compareExternals(Path1, Path2, resultFile):
    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # call SVN to get externals
    #analyzePath(sys.argv[1])
    p1 = analyzePath(Path1)
    if Path2:
        p2 = analyzePath(Path2)

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # summarize data
    worksheet_lines = list()
    for external in p1:
        line = list()
        line.append(external[EXT_NAME])
        line.append(external[EXT_PATH])
        line.append(external[EXT_REVISION])
        worksheet_lines.append(line)
    if Path2:
        for external in p2:
            try:
                existingInP1 = next(line for line in worksheet_lines if line[0]==external[EXT_NAME])
                existingInP1.append(external[EXT_PATH])
                existingInP1.append(external[EXT_REVISION])
            except:
                # new line
                worksheet_lines.append([external[EXT_NAME],
                                        "",
                                        "",
                                        external[EXT_PATH],
                                        external[EXT_REVISION]])

    #~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
    # generate Excel Workbook
    wb = Workbook()
    ws = wb.active # get active worksheet

    # header
    ws.append(["external name", Path1, "revision", Path2, "revision"])

    # set column widths
    for i, width in enumerate([30,100,12,100,12]):
        ws.column_dimensions[get_column_letter(i+1)].width = width

    # write data to worksheet
    for line in worksheet_lines:
        ws.append(line)

    # activate Auto Filter
    ws.sheet_properties.filterMode = True

    # highlight the differences
    bold = Font(bold=True)
    for cell in ['A1','B1','C1','D1','E1']:
        ws[cell].font=bold

    highlight = Font(color="FF0000")
    for row in ws.iter_rows(min_row=2):
        if (row[1].value != row[3].value):
            row[3].font = highlight
        if (row[2].value != row[4].value):
            row[4].font = highlight

    # Save the file
    wb.save(resultFile)

if __name__ == '__main__':
    # Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PlanningAssistant/trunk"
    # Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PlanningAssistant/branches/RefractiveWorkplace"
    # resultFile = "comparison_PA_NeoTrunk2RefWP.xlsx"
    # compareExternals(Path1, Path2, resultFile)
    #
    # Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PlanningAssistant/trunk/"
    # Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/CommonComponents/PlanningAssistantComponents/ExcimerPlanning/trunk"
    # resultFile = "comparison_PA_NeoTrunk2EP.xlsx"
    # compareExternals(Path1, Path2, resultFile)
    #
    # Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PlanningAssistant/trunk"
    # Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PlanningAssistant/branches/releases/1.0"
    # resultFile = "comparison_PA_NeoTrunk2Release10.xlsx"
    # compareExternals(Path1, Path2, resultFile)
    #
    # Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PlanningAssistant/tags/1.0.0_NEO_RC3"
    # Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PlanningAssistant/branches/RefractiveWorkplace"
    # resultFile = "comparison_PA_Tag_NEO_RC3_vs_REFWP_trunk.xlsx"
    # compareExternals(Path1, Path2, resultFile)
    #
    # Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/CommonLocalization/branches/releases/1.0"
    # Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/RefractiveWorkplace/RefWPLocalization"
    # resultFile = "externals_CommonLocalization_Neo_vs_REFWPLocalization.xlsx"
    # compareExternals(Path1, Path2, resultFile)

    #Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PlanningAssistant/tags/1.0.0_RefWP_RC2"
    #Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PlanningAssistant/tags/1.0.0_NEO_RC5"
    #resultFile = "comparison_PA_NEO_1.0.0RC5_vs_RefWP_1.0.0.RC2.xlsx"
    #compareExternals(Path1, Path2, resultFile)

    Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/RefractiveWorkplace/RefWPClientInstaller/branches/releases/1.1"
    Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Setup/SystemInstaller/tags/1.1.0_NEO_RC7_PackageVersion"
    resultFile = "comparison_installer_NEO_tag11_RC7_vs_RefWP_r11.xlsx"
    compareExternals(Path1, Path2, resultFile)

    # Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/RefractiveWorkplace/RefWPClientInstaller/branches/releases/1.1"
    # Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/RefractiveWorkplace/RefWPClientInstaller/tags/RefWP_1.1.0_RC2"
    # resultFile = "comparison_installer_RefWP_r11_RC2_vs_RefWP_branch_r11.xlsx"
    # compareExternals(Path1, Path2, resultFile)

    # Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/CommonComponents/PlanningAssistantComponents/ExcimerPlanning/branches/dev/ExcimerPlanningPRK1.0"
    # Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/CommonComponents/PlanningAssistantComponents/ExcimerPlanning/branches/rc-1.0.0"
    # resultFile = "comparison_PRK1.0_vs_EP_rc100.xlsx"
    # compareExternals(Path1, Path2, resultFile)

    # Path1 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PatternProcessing/branches/releases/1.0"
    # Path2 = "https://gullstrand.zeiss.org/svn/czm/RefractiveLasers/NEO/Applications/PatternProcessing/tags/1.0.0_NEO_RC4"
    # resultFile = "externals_PatternProcessing_R1.0_vs_tag1.0.0_RC4.xlsx"
    # compareExternals(Path1, Path2, resultFile)
